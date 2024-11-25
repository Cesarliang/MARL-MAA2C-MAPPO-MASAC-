from __future__ import print_function, division

import time

from openpyxl.reader.excel import load_workbook
from tqdm import tqdm

from MAA2C import MAA2C
from common.utils import agg_double_list, copy_file, init_dir
from datetime import datetime

import argparse
import configparser
import sys
sys.path.append("../highway-env")

import gym
import os
import highway_env
import numpy as np
import matplotlib.pyplot as plt
path = 'F:\新疆大学\强化学习\大论文实验汇总\MARL-meta\测试\学习率a调整\MAA2C-HARD-1000-a5e-10.xlsx'
# path2 = "F:\新疆大学\强化学习\\多智能体\MAA2C2.xlsx"
# path = '/root/autodl-tmp/MARL/MAPPO1.xlsx'
# path2 ='/root/autodl-tmp/MARL/MAPPO2.xlsx'
workbook = load_workbook(path)
# workbook = load_workbook(path)
# workbook2 = load_workbook(path2)
sheet = workbook['Sheet1']
# sheet2 = workbook2['Sheet1']

def parse_args():#解析命令行参数的函数
    """
    Description for this experiment:
        + hard: 7-steps, curriculum
    """
    default_base_dir = "./results/"
    default_config_dir = 'configs/configs.ini'
    parser = argparse.ArgumentParser(description=('Train or evaluate policy on RL environment '
                                                  'using MA2C'))
    parser.add_argument('--base-dir', type=str, required=False,
                        default=default_base_dir, help="experiment base dir")
    parser.add_argument('--option', type=str, required=False,
                        default='train', help="train or evaluate")
    parser.add_argument('--config-dir', type=str, required=False,
                        default=default_config_dir, help="experiment config path")
    parser.add_argument('--model-dir', type=str, required=False,
                        default='', help="pretrained model path")
    parser.add_argument('--evaluation-seeds', type=str, required=False,
                        default=','.join([str(i) for i in range(0, 600, 20)]),
                        help="random seeds for evaluation, split by ,")
    args = parser.parse_args()
    return args
#以用于在训练或评估模式下设置实验的基本目录、配置文件路径、预训练模型路径以及评估所使用的随机种子。

def train(args):
    base_dir = args.base_dir
    config_dir = args.config_dir
    config = configparser.ConfigParser()
    config.read(config_dir)

    # create an experiment folder
    now = datetime.now().strftime("%b-%d_%H_%M_%S")
    output_dir = base_dir + now
    dirs = init_dir(output_dir)
    copy_file(dirs['configs'])

    if os.path.exists(args.model_dir):
        model_dir = args.model_dir
    else:
        model_dir = dirs['models']

    # model configs 获取了一系列的模型配置参数。根据配置文件的内容
    BATCH_SIZE = config.getint('MODEL_CONFIG', 'BATCH_SIZE') #批次大小
    MEMORY_CAPACITY = config.getint('MODEL_CONFIG', 'MEMORY_CAPACITY')#经验回放缓冲区的容量
    ROLL_OUT_N_STEPS = config.getint('MODEL_CONFIG', 'ROLL_OUT_N_STEPS')#多步回放的步数
    reward_gamma = config.getfloat('MODEL_CONFIG', 'reward_gamma')
    training_strategy = config.get('MODEL_CONFIG', 'training_strategy')
    actor_hidden_size = config.getint('MODEL_CONFIG', 'actor_hidden_size')
    critic_hidden_size = config.getint('MODEL_CONFIG', 'critic_hidden_size')
    MAX_GRAD_NORM = config.getfloat('MODEL_CONFIG', 'MAX_GRAD_NORM')#梯度裁剪的最大范数
    ENTROPY_REG = config.getfloat('MODEL_CONFIG', 'ENTROPY_REG')#熵正则化的权重
    epsilon = config.getfloat('MODEL_CONFIG', 'epsilon')
    alpha = config.getfloat('MODEL_CONFIG', 'alpha')
    state_split = config.getboolean('MODEL_CONFIG', 'state_split')
    shared_network = config.getboolean('MODEL_CONFIG', 'shared_network')
    reward_type = config.get('MODEL_CONFIG', 'reward_type')

    # train configs 训练配置
    actor_lr = config.getfloat('TRAIN_CONFIG', 'actor_lr')
    critic_lr = config.getfloat('TRAIN_CONFIG', 'critic_lr')
    MAX_EPISODES = config.getint('TRAIN_CONFIG', 'MAX_EPISODES')
    EPISODES_BEFORE_TRAIN = config.getint('TRAIN_CONFIG', 'EPISODES_BEFORE_TRAIN')#在开始训练之前的回合数（用于预热经验回放缓冲区）
    EVAL_INTERVAL = config.getint('TRAIN_CONFIG', 'EVAL_INTERVAL') #评估模型的间隔（以训练回合数为单位）
    EVAL_EPISODES = config.getint('TRAIN_CONFIG', 'EVAL_EPISODES')#每次评估的回合数
    reward_scale = config.getfloat('TRAIN_CONFIG', 'reward_scale')#奖励缩放因子

    # init env
    env = gym.make('merge-multi-agent-v0')
    env.config['seed'] = config.getint('ENV_CONFIG', 'seed')#种子
    env.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')#模拟频率
    env.config['duration'] = config.getint('ENV_CONFIG', 'duration')#持续时间
    env.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')#策略更新频率
    env.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')#碰撞奖励
    env.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')#高速奖励
    env.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')#车头时局成本
    env.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')#
    env.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')#汇入成本
    env.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')#交通密度
    env.config['safety_guarantee'] = config.getboolean('ENV_CONFIG', 'safety_guarantee')
    env.config['n_step'] = config.getint('ENV_CONFIG', 'n_step')
    traffic_density = config.getint('ENV_CONFIG', 'traffic_density')
    env.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    assert env.T % ROLL_OUT_N_STEPS == 0#用断言确保环境的总步数（env.T）能够被ROLL_OUT_N_STEPS整除，以确保正确的步长设置。

    env_eval = gym.make('merge-multi-agent-v0')# 创建了  env_eval的环境 和上面env环境相同
    env_eval.config['seed'] = config.getint('ENV_CONFIG', 'seed') + 1
    env_eval.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env_eval.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env_eval.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env_eval.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env_eval.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env_eval.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env_eval.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env_eval.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env_eval.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    env_eval.config['safety_guarantee'] = config.getboolean('ENV_CONFIG', 'safety_guarantee')
    env_eval.config['n_step'] = config.getint('ENV_CONFIG', 'n_step')
    env_eval.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    state_dim = env.n_s
    action_dim = env.n_a
    test_seeds = args.evaluation_seeds

    ma2c = MAA2C(env, state_dim=state_dim, action_dim=action_dim,
                 memory_capacity=MEMORY_CAPACITY, max_steps=None,
                 roll_out_n_steps=ROLL_OUT_N_STEPS,
                 reward_gamma=reward_gamma, reward_scale=reward_scale, done_penalty=None,
                 actor_hidden_size=actor_hidden_size, critic_hidden_size=critic_hidden_size,
                 actor_lr=actor_lr, critic_lr=critic_lr,
                 optimizer_type="rmsprop", entropy_reg=ENTROPY_REG,
                 max_grad_norm=MAX_GRAD_NORM, batch_size=BATCH_SIZE,
                 episodes_before_train=EPISODES_BEFORE_TRAIN,
                 use_cuda=False, training_strategy=training_strategy,
                 epsilon=epsilon, alpha=alpha, traffic_density=traffic_density, test_seeds=test_seeds,
                 state_split=state_split, shared_network=shared_network, reward_type=reward_type)

    # load the model if exist
    ma2c.load(model_dir, train_mode=True)#加载预训练模型（如果存在），其中model_dir为预训练模型的路径。
    env.seed = env.config['seed']#设置环境的随机种子，以确保可复现性。
    env.unwrapped.seed = env.config['seed']
    # print(env.seed)
    episodes = []
    eval_rewards = []
    best_eval_reward = -100
    ep_rewards = []
    speeds = []
    v_losss = []
    i = 0
    j = 1
    cursh_num = 0
    sheet.cell(1, 2, "回合奖励")
    sheet.cell(1, 3, "回合损失")
    sheet.cell(1, 4, "回合速度")
    sheet.cell(1, 6, "评价奖励")
    sheet.cell(1, 7, "平均速度")
    sheet.cell(1, 8, "碰撞次数")
    sheet.cell(1, 9, "碰撞率")
    sheet.cell(1, 10, "车头时距")
    sheet.cell(1, 11, "汇入成本")
    print("路径",path,"模式",ma2c.traffic_density,"seeds",env.seed,"EVAL_INTERVAL",EVAL_INTERVAL)
    print("share-net",ma2c.shared_network,"类别",ma2c.training_strategy)
    while ma2c.n_episodes < MAX_EPISODES:
        ma2c.explore()
        if ma2c.n_episodes >= EPISODES_BEFORE_TRAIN:
            ma2c.train()
        # print(ma2c.episode_rewards,'22')
        i = ma2c.n_episodes - 1
        # if i ==1:
        #     print(ma2c.actors)
        reward = float(ma2c.episode_rewards[i])
        speed = float(ma2c.average_speed[i])

        cursh_num += ma2c.cursh
        # print(reward,type(reward))
        print("回合数：", i, "当前回合奖励：", reward,  "平均速度", speed,"车辆数量",ma2c.n_agents)
        ep_rewards.append(reward)

        speeds.append(speed)
        v_losss.append(ma2c.v_loss1)
        if ((i % 200) == 0 or i == 10) and i != 0:
            qidian = 0
            now = time.localtime()
            nowt = time.strftime("%Y-%m-%d/%H.%M.%S", now)
            print('列入中,当前回合：', i, '时间：', nowt)
            if i % 200 == 0:
                qidian = i - 200
            for k in tqdm(range(qidian, i)):
                time.sleep(0.1)
                sheet.cell(k + 2, 1, k)

                # print(rewards[0],rewards[k])
                sheet.cell(k + 2, 2, ep_rewards[k])
                sheet.cell(k + 2, 3, v_losss[k])
                sheet.cell(k + 2, 4, speeds[k])
            workbook.save(path)


        if ma2c.episode_done and ((ma2c.n_episodes + 1) % EVAL_INTERVAL == 0):#检查当回合是否结束并达到评价的间隔
            j = j + 1
            rewards, _, _, speeds_eval,headway_dis,merging_cost = ma2c.evaluation(env_eval, dirs['train_videos'], EVAL_EPISODES)
            rewards_mu, rewards_std = agg_double_list(rewards)
            # print("Episode %d, Average Reward %.2f" % (ma2c.n_episodes + 1, rewards_mu))
            episodes.append(ma2c.n_episodes + 1)
            av_speed = np.mean(speeds_eval)
            av_head_time = np.mean(headway_dis) / speed
            av_merging_cost = np.mean(merging_cost)
            # print(rewards_std,speeds,av_speed,111)
            print('回合', ma2c.n_episodes + 1, "平均奖励", rewards_mu, "奖励标准差", rewards_std, '平均速度', av_speed,
                  "平均车头时距", av_head_time, "汇入成本", av_merging_cost)
            print("路径", path)
            eval_rewards.append(rewards_mu)

            # save the model
            # if rewards_mu > best_eval_reward:
            #     ma2c.save(dirs['models'], 100000)
            #     ma2c.save(dirs['models'], ma2c.n_episodes + 1)
            #     best_eval_reward = rewards_mu#如果当前评估奖励rewards_mu超过了之前的最佳评估奖励
            #     # best_eval_reward，则保存模型为当前回合
            # else:
            #     ma2c.save(dirs['models'], ma2c.n_episodes + 1)
            print("eval_reward导入")
            sheet.cell(j, 6, float(rewards_mu))
            # sheet.cell(j, 9, float(rewards_std))
            sheet.cell(j, 7, float(av_speed))
            sheet.cell(j, 8, float(cursh_num))
            sheet.cell(j, 9, float(cursh_num) / 200)
            sheet.cell(j, 10, float(av_head_time))
            sheet.cell(j, 11, float(av_merging_cost))
            workbook.save(path)
            cursh_num = 0
        # np.save(output_dir + '/{}'.format('eval_rewards'), np.array(eval_rewards))
        # # save training data
        # np.save(output_dir + '/{}'.format('episode_rewards'), np.array(ma2c.episode_rewards))
        # np.save(output_dir + '/{}'.format('epoch_steps'), np.array(ma2c.epoch_steps))
        # np.save(output_dir + '/{}'.format('average_speed'), np.array(ma2c.average_speed))
#!!!这部分可以调整成表格形式
    # save the model
    ma2c.save(dirs['models'], MAX_EPISODES + 2)#保存模型，使用+2为全局步数，用于最终保存的模型
    #探索完才画图？
    plt.figure()
    plt.plot(eval_rewards)
    plt.xlabel("Episode")
    plt.ylabel("Average Training Reward")
    plt.legend(["MAA2C"]) # 画图？
    plt.savefig(output_dir + '/' + "maa2c_train.png")
    plt.show()


def evaluate(args): #评价函数
    if os.path.exists(args.model_dir):
        model_dir = args.model_dir + '/models/'
    else:
        raise Exception("Sorry, no pretrained models")
    config_dir = args.model_dir + '/configs/configs.ini'
    config = configparser.ConfigParser()
    config.read(config_dir)

    video_dir = args.model_dir + '/eval_videos'
    eval_logs = args.model_dir + '/eval_logs'

    # model configs
    BATCH_SIZE = config.getint('MODEL_CONFIG', 'BATCH_SIZE')
    MEMORY_CAPACITY = config.getint('MODEL_CONFIG', 'MEMORY_CAPACITY')
    ROLL_OUT_N_STEPS = config.getint('MODEL_CONFIG', 'ROLL_OUT_N_STEPS')
    reward_gamma = config.getfloat('MODEL_CONFIG', 'reward_gamma')
    training_strategy = config.get('MODEL_CONFIG', 'training_strategy')
    actor_hidden_size = config.getint('MODEL_CONFIG', 'actor_hidden_size')
    critic_hidden_size = config.getint('MODEL_CONFIG', 'critic_hidden_size')
    MAX_GRAD_NORM = config.getfloat('MODEL_CONFIG', 'MAX_GRAD_NORM')
    ENTROPY_REG = config.getfloat('MODEL_CONFIG', 'ENTROPY_REG')
    epsilon = config.getfloat('MODEL_CONFIG', 'epsilon')
    alpha = config.getfloat('MODEL_CONFIG', 'alpha')
    state_split = config.getboolean('MODEL_CONFIG', 'state_split')
    shared_network = config.getboolean('MODEL_CONFIG', 'shared_network')
    reward_type = config.get('MODEL_CONFIG', 'reward_type')

    # train configs
    actor_lr = config.getfloat('TRAIN_CONFIG', 'actor_lr')
    critic_lr = config.getfloat('TRAIN_CONFIG', 'critic_lr')
    EPISODES_BEFORE_TRAIN = config.getint('TRAIN_CONFIG', 'EPISODES_BEFORE_TRAIN')
    reward_scale = config.getfloat('TRAIN_CONFIG', 'reward_scale')

    # init env
    env = gym.make('merge-multi-agent-v0')
    env.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    env.config['safety_guarantee'] = config.getboolean('ENV_CONFIG', 'safety_guarantee')
    env.config['n_step'] = config.getint('ENV_CONFIG', 'n_step')
    traffic_density = config.getint('ENV_CONFIG', 'traffic_density')
    env.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    assert env.T % ROLL_OUT_N_STEPS == 0
    state_dim = env.n_s
    action_dim = env.n_a
    test_seeds = args.evaluation_seeds
    seeds = [int(s) for s in test_seeds.split(',')]

    ma2c = MAA2C(env, state_dim=state_dim, action_dim=action_dim,
                 memory_capacity=MEMORY_CAPACITY, max_steps=None,
                 roll_out_n_steps=ROLL_OUT_N_STEPS,
                 reward_gamma=reward_gamma, reward_scale=reward_scale, done_penalty=None,
                 actor_hidden_size=actor_hidden_size, critic_hidden_size=critic_hidden_size,
                 actor_lr=actor_lr, critic_lr=critic_lr,
                 optimizer_type="rmsprop", entropy_reg=ENTROPY_REG,
                 max_grad_norm=MAX_GRAD_NORM, batch_size=BATCH_SIZE,
                 episodes_before_train=EPISODES_BEFORE_TRAIN,
                 use_cuda=False, training_strategy=training_strategy,
                 epsilon=epsilon, alpha=alpha, traffic_density=traffic_density, test_seeds=test_seeds,
                 state_split=state_split, shared_network=shared_network, reward_type=reward_type)

    # load the model if exist 加载模型
    ma2c.load(model_dir, train_mode=False)#加载预训练模型 准备进行评估
    rewards, (vehicle_speed, vehicle_position), steps, avg_speeds = ma2c.evaluation(env, video_dir, len(seeds), is_train=False)
    rewards_mu, rewards_std = agg_double_list(rewards)#奖励的均值和标准差
    success_rate = sum(np.array(steps) == 100) / len(steps)#完成100步的回合数占总回合数
    avg_speeds_mu, avg_speeds_std = agg_double_list(avg_speeds)

    print("Evaluation Reward and std %.2f, %.2f " % (rewards_mu, rewards_std))
    print("Collision Rate %.2f" % (1 - success_rate))
    print("Average Speed and std %.2f , %.2f " % (avg_speeds_mu, avg_speeds_std))

    np.save(eval_logs + '/{}'.format('eval_rewards'), np.array(rewards))
    np.save(eval_logs + '/{}'.format('eval_steps'), np.array(steps))
    np.save(eval_logs + '/{}'.format('eval_avg_speeds'), np.array(avg_speeds))
    np.save(eval_logs + '/{}'.format('vehicle_speed'), np.array(vehicle_speed))
    np.save(eval_logs + '/{}'.format('vehicle_position'), np.array(vehicle_position))


if __name__ == "__main__":
    args = parse_args()
    # train or eval
    if args.option == 'train':
        train(args)
    else:
        evaluate(args)
