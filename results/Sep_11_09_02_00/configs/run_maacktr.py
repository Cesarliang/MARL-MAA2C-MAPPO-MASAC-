import time

from openpyxl.reader.excel import load_workbook
from tqdm import tqdm

from MAACKTR import JointACKTR as MAACKTR
from common.utils import agg_double_list, copy_file_akctr, init_dir

import sys
sys.path.append("../highway-env")
import os
import gym
import numpy as np
import matplotlib.pyplot as plt
import highway_env
from datetime import datetime

import argparse
import configparser

path = 'F:\新疆大学\强化学习\\多智能体\MAACKTR1.xlsx'
path2 = "F:\新疆大学\强化学习\\多智能体\MAACKTR2.xlsx"
# path = '/root/autodl-tmp/MARL/MAPPO1.xlsx'
# path2 ='/root/autodl-tmp/MARL/MAPPO2.xlsx'
workbook = load_workbook(path)
# workbook = load_workbook(path)
workbook2 = load_workbook(path2)
sheet = workbook['Sheet1']
sheet2 = workbook2['Sheet1']
def parse_args():
    """
    Description for this experiment:
        + medium: maacktr, regionalR
        + seed = 0
    """
    default_base_dir = "./results/"
    default_config_dir = 'configs/configs_acktr.ini'
    parser = argparse.ArgumentParser(description=('Train or evaluate policy on RL environment '
                                                  'using maacktr'))
    parser.add_argument('--base-dir', type=str, required=False,
                        default=default_base_dir, help="experiment base dir")
    parser.add_argument('--option', type=str, required=False,
                        default='train', help="train or evaluate")
    parser.add_argument('--config-dir', type=str, required=False,
                        default=default_config_dir, help="experiment config path")
    parser.add_argument('--model-dir', type=str, required=False,
                        default='', help="pretrained model path")
    parser.add_argument('--evaluation-seeds', type=str, required=False,
                        default=','.join([str(i) for i in range(0, 600, 20)]),
                        help="random seeds for evaluation, split by ,")
    args = parser.parse_args()
    return args


def train(args):
    base_dir = args.base_dir
    config_dir = args.config_dir
    config = configparser.ConfigParser()
    config.read(config_dir)

    # create an experiment folder
    now = datetime.utcnow().strftime("%b_%d_%H_%M_%S")
    output_dir = base_dir + now
    dirs = init_dir(output_dir)
    copy_file_akctr(dirs['configs'])

    if os.path.exists(args.model_dir):
        model_dir = args.model_dir
    else:
        model_dir = dirs['models']

    # model configs
    BATCH_SIZE = config.getint('MODEL_CONFIG', 'BATCH_SIZE')
    MEMORY_CAPACITY = config.getint('MODEL_CONFIG', 'MEMORY_CAPACITY')
    ROLL_OUT_N_STEPS = config.getint('MODEL_CONFIG', 'ROLL_OUT_N_STEPS')
    reward_gamma = config.getfloat('MODEL_CONFIG', 'reward_gamma')
    actor_hidden_size = config.getint('MODEL_CONFIG', 'actor_hidden_size')
    critic_hidden_size = config.getint('MODEL_CONFIG', 'critic_hidden_size')
    MAX_GRAD_NORM = config.getfloat('MODEL_CONFIG', 'MAX_GRAD_NORM')
    ENTROPY_REG = config.getfloat('MODEL_CONFIG', 'ENTROPY_REG')
    reward_type = config.get('MODEL_CONFIG', 'reward_type')

    # train configs
    actor_lr = config.getfloat('TRAIN_CONFIG', 'actor_lr')
    critic_lr = config.getfloat('TRAIN_CONFIG', 'critic_lr')
    MAX_EPISODES = config.getint('TRAIN_CONFIG', 'MAX_EPISODES')
    EPISODES_BEFORE_TRAIN = config.getint('TRAIN_CONFIG', 'EPISODES_BEFORE_TRAIN')
    EVAL_INTERVAL = config.getint('TRAIN_CONFIG', 'EVAL_INTERVAL')
    EVAL_EPISODES = config.getint('TRAIN_CONFIG', 'EVAL_EPISODES')
    reward_scale = config.getfloat('TRAIN_CONFIG', 'reward_scale')

    # init env
    env = gym.make('merge-multi-agent-v0')
    env.config['seed'] = config.getint('ENV_CONFIG', 'seed')
    env.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    traffic_density = config.getint('ENV_CONFIG', 'traffic_density')
    env.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    assert env.T % ROLL_OUT_N_STEPS == 0

    env_eval = gym.make('merge-multi-agent-v0')
    env_eval.config['seed'] = config.getint('ENV_CONFIG', 'seed') + 1
    env_eval.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env_eval.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env_eval.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env_eval.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env_eval.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env_eval.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env_eval.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env_eval.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env_eval.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    env_eval.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    state_dim = env.n_s
    action_dim = env.n_a
    test_seeds = args.evaluation_seeds
    maacktr = MAACKTR(env=env, memory_capacity=MEMORY_CAPACITY,
                      state_dim=state_dim, action_dim=action_dim,
                      batch_size=BATCH_SIZE, entropy_reg=ENTROPY_REG,
                      actor_lr=actor_lr, critic_lr=critic_lr,
                      reward_gamma=reward_gamma, reward_scale=reward_scale,
                      actor_hidden_size=actor_hidden_size, critic_hidden_size=critic_hidden_size,
                      roll_out_n_steps=ROLL_OUT_N_STEPS, test_seeds=test_seeds,
                      max_grad_norm=MAX_GRAD_NORM, reward_type=reward_type,
                      episodes_before_train=EPISODES_BEFORE_TRAIN, traffic_density=traffic_density)

    # load the model if exist
    maacktr.load(model_dir, train_mode=True)
    env.seed = env.config['seed']
    env.unwrapped.seed = env.config['seed']
    eval_rewards = []
    print("train begin")

    ep_rewards = []
    speeds = []
    v_losss = []
    i = 0
    j = 0
    while maacktr.n_episodes < MAX_EPISODES:
        maacktr.interact()
        if maacktr.n_episodes >= EPISODES_BEFORE_TRAIN:
            maacktr.train()
        i = maacktr.n_episodes - 1
        reward = float(maacktr.episode_rewards[i])
        speed = float(maacktr.average_speed[i])
        # print(reward,type(reward))
        print("回合数：", i, "当前回合奖励：", reward
              , "v_loss:", maacktr.v_loss1, "平均速度", speed)
        ep_rewards.append(reward)
        speeds.append(speed)
        v_losss.append(maacktr.v_loss1)
        if ((i % 15) == 0 or i == 10) and i != 0:
            qidian = 0
            now = time.localtime()
            nowt = time.strftime("%Y-%m-%d/%H.%M.%S", now)
            print('列入中,当前回合：', i, '时间：', nowt)
            if i % 15 == 0:
                qidian = i - 15
            for k in tqdm(range(qidian, i)):
                time.sleep(0.1)
                sheet.cell(k + 2, 1, k)

                # print(rewards[0],rewards[k])
                sheet.cell(k + 2, 2, ep_rewards[k])
                sheet.cell(k + 2, 4, v_losss[k])
                sheet.cell(k + 2, 6, speeds[k])
            workbook.save(path)
        if maacktr.episode_done and ((maacktr.n_episodes + 1) % EVAL_INTERVAL == 0):
            rewards, _, _, _ = maacktr.evaluation(env_eval, dirs['train_videos'], EVAL_EPISODES)
            rewards_mu, rewards_std = agg_double_list(rewards)
            j = j +1
            print("Episode %d, Average Reward %.2f" % (maacktr.n_episodes + 1, rewards_mu))
            eval_rewards.append(rewards_mu)
            # save the model
            maacktr.save(dirs['models'], maacktr.n_episodes + 1)
            print("eval_reward导入")
            sheet2.cell(j, 1, float(rewards_mu))
            workbook2.save(path2)
    # save the model
    maacktr.save(dirs['models'], MAX_EPISODES + 2)
    plt.figure()
    plt.plot(eval_rewards)
    plt.xlabel("Episode")
    plt.ylabel("Average Reward")
    plt.legend(["MAACKTR"])
    plt.show()


def evaluate(args):
    if os.path.exists(args.model_dir):
        model_dir = args.model_dir + '/models/'
    else:
        raise Exception("Sorry, no pretrained models")
    config_dir = args.model_dir + '/configs/configs_acktr.ini'
    config = configparser.ConfigParser()
    config.read(config_dir)

    video_dir = args.model_dir + '/eval_videos'

    # model configs
    BATCH_SIZE = config.getint('MODEL_CONFIG', 'BATCH_SIZE')
    MEMORY_CAPACITY = config.getint('MODEL_CONFIG', 'MEMORY_CAPACITY')
    ROLL_OUT_N_STEPS = config.getint('MODEL_CONFIG', 'ROLL_OUT_N_STEPS')
    reward_gamma = config.getfloat('MODEL_CONFIG', 'reward_gamma')
    actor_hidden_size = config.getint('MODEL_CONFIG', 'actor_hidden_size')
    critic_hidden_size = config.getint('MODEL_CONFIG', 'critic_hidden_size')
    MAX_GRAD_NORM = config.getfloat('MODEL_CONFIG', 'MAX_GRAD_NORM')
    ENTROPY_REG = config.getfloat('MODEL_CONFIG', 'ENTROPY_REG')
    reward_type = config.get('MODEL_CONFIG', 'reward_type')

    # train configs
    actor_lr = config.getfloat('TRAIN_CONFIG', 'actor_lr')
    critic_lr = config.getfloat('TRAIN_CONFIG', 'critic_lr')
    EPISODES_BEFORE_TRAIN = config.getint('TRAIN_CONFIG', 'EPISODES_BEFORE_TRAIN')
    reward_scale = config.getfloat('TRAIN_CONFIG', 'reward_scale')

    # init env
    env = gym.make('merge-multi-agent-v0')
    env.config['seed'] = config.getint('ENV_CONFIG', 'seed')
    env.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    traffic_density = config.getint('ENV_CONFIG', 'traffic_density')
    env.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    assert env.T % ROLL_OUT_N_STEPS == 0
    state_dim = env.n_s
    action_dim = env.n_a
    test_seeds = args.evaluation_seeds
    seeds = [int(s) for s in test_seeds.split(',')]

    maacktr = MAACKTR(env=env, memory_capacity=MEMORY_CAPACITY,
                      state_dim=state_dim, action_dim=action_dim,
                      batch_size=BATCH_SIZE, entropy_reg=ENTROPY_REG,
                      actor_lr=actor_lr, critic_lr=critic_lr,
                      reward_gamma=reward_gamma, reward_scale=reward_scale,
                      actor_hidden_size=actor_hidden_size, critic_hidden_size=critic_hidden_size,
                      roll_out_n_steps=ROLL_OUT_N_STEPS, test_seeds=test_seeds,
                      max_grad_norm=MAX_GRAD_NORM, reward_type=reward_type,
                      episodes_before_train=EPISODES_BEFORE_TRAIN, traffic_density=traffic_density)

    # load the model if exist
    maacktr.load(model_dir, train_mode=False)
    rewards, _, steps, avg_speeds = maacktr.evaluation(env, video_dir, len(seeds), is_train=False)


if __name__ == "__main__":
    args = parse_args()
    # train or eval
    if args.option == 'train':
        train(args)
    else:
        evaluate(args)
