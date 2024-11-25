import time

from numpy import mean

from MASAC import (MASAC)
from server import SERVER
from common.utils import agg_double_list, copy_file_sac, init_dir
import sys
sys.path.append("../highway-env")
from tqdm import tqdm
import gym
import numpy as np
import matplotlib.pyplot as plt
import highway_env
import argparse
import configparser
import os
from datetime import datetime
from openpyxl import load_workbook
import bisect
# path = 'E:\新疆大学\强化学习\\45-test-jump.xlsx'
# path = 'E:\新疆大学\强化学习\多智能体\\bendi\PPO-MEDIUM_0fenbu_avg.xlsx'
# path2 = "E:\新疆大学\强化学习\\多智能体\MAPPO2.xlsx"
path = 'E:\李春\\FMMASAC.xlsx'
# path2 ='/root/autodl-tmp/MARL/MAPPO2.xlsx'
workbook = load_workbook(path)
# workbook = load_workbook(path)
# workbook2 = load_workbook(path2)
sheet = workbook['Sheet1']
# sheet2 = workbook2['Sheet1']
# sheet = workbook['Sheet1']
import warnings
warnings.filterwarnings("ignore")

def parse_args():
    """
    Description for this experiment:
        + easy: globalR
        + seed = 0
    """
    default_base_dir = "./results/"
    default_config_dir = 'configs/configs_sac.ini'
    parser = argparse.ArgumentParser(description=('Train or evaluate policy on RL environment '
                                                  'using masac'))
    parser.add_argument('--base-dir', type=str, required=False,
                        default=default_base_dir, help="experiment base dir")
    parser.add_argument('--option', type=str, required=False,
                        default='train', help="train or evaluate")
    parser.add_argument('--config-dir', type=str, required=False,
                        default=default_config_dir, help="experiment config path")
    parser.add_argument('--model-dir', type=str, required=False,
                        default='', help="pretrained model path")
    parser.add_argument('--evaluation-seeds', type=str, required=False,
                        default=','.join([str(i) for i in range(0, 600, 20)]),
                        help="random seeds for evaluation, split by ,")
    parser.add_argument('--hidden_dim', type=int, default=128, help='Dimension of hidden layers')
    args = parser.parse_args()
    return args

def train(args):
    hidden_dim = args.hidden_dim
    base_dir = args.base_dir #这里可以修改配置文件
    config_dir = args.config_dir
    config = configparser.ConfigParser()
    config.read(config_dir)

    # create an experiment folder
    now = datetime.utcnow().strftime("%b_%d_%H_%M_%S")
    output_dir = base_dir + now
    dirs = init_dir(output_dir)
    copy_file_sac(dirs['configs'])

    if os.path.exists(args.model_dir):
        model_dir = args.model_dir
    else:
        model_dir = dirs['models']

    # model configs
    BATCH_SIZE = config.getint('MODEL_CONFIG', 'BATCH_SIZE')
    MEMORY_CAPACITY = config.getint('MODEL_CONFIG', 'MEMORY_CAPACITY')
    ROLL_OUT_N_STEPS = config.getint('MODEL_CONFIG', 'ROLL_OUT_N_STEPS')
    reward_gamma = config.getfloat('MODEL_CONFIG', 'reward_gamma')
    actor_hidden_size = config.getint('MODEL_CONFIG', 'actor_hidden_size')
    critic_hidden_size = config.getint('MODEL_CONFIG', 'critic_hidden_size')
    MAX_GRAD_NORM = config.getfloat('MODEL_CONFIG', 'MAX_GRAD_NORM')
    ENTROPY_REG = config.getfloat('MODEL_CONFIG', 'ENTROPY_REG')
    reward_type = config.get('MODEL_CONFIG', 'reward_type')
    TARGET_UPDATE_STEPS = config.getint('MODEL_CONFIG', 'TARGET_UPDATE_STEPS')
    TARGET_TAU = config.getfloat('MODEL_CONFIG', 'TARGET_TAU')

    # train configs
    actor_lr = config.getfloat('TRAIN_CONFIG', 'actor_lr')
    critic_lr = config.getfloat('TRAIN_CONFIG', 'critic_lr')
    MAX_EPISODES = config.getint('TRAIN_CONFIG', 'MAX_EPISODES')
    EPISODES_BEFORE_TRAIN = config.getint('TRAIN_CONFIG', 'EPISODES_BEFORE_TRAIN')
    EVAL_INTERVAL = config.getint('TRAIN_CONFIG', 'EVAL_INTERVAL')
    EVAL_EPISODES = config.getint('TRAIN_CONFIG', 'EVAL_EPISODES')
    reward_scale = config.getfloat('TRAIN_CONFIG', 'reward_scale')

    # init env
    env = gym.make('merge-multi-agent-v0')
    env.config['seed'] = config.getint('ENV_CONFIG', 'seed')
    env.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    traffic_density = config.getint('ENV_CONFIG', 'traffic_density')
    env.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    assert env.T % ROLL_OUT_N_STEPS == 0

    env_eval = gym.make('merge-multi-agent-v0')
    env_eval.config['seed'] = config.getint('ENV_CONFIG', 'seed') + 1
    env_eval.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env_eval.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env_eval.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env_eval.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env_eval.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env_eval.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env_eval.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env_eval.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env_eval.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    env_eval.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    state_dim = env.n_s
    action_dim = env.n_a
    test_seeds = args.evaluation_seeds
    n_agents = len(env.controlled_vehicles)

    masac = MASAC(env=env, n_agents=n_agents, memory_capacity=MEMORY_CAPACITY,
                  state_dim=state_dim, action_dim=action_dim, hidden_dim=hidden_dim,
                  batch_size=BATCH_SIZE, actor_hidden_dim=actor_hidden_size, critic_hidden_dim=critic_hidden_size,
                  actor_lr=actor_lr, critic_lr=critic_lr, reward_scale=reward_scale, max_grad_norm=MAX_GRAD_NORM,
                  entropy_reg=ENTROPY_REG, tau=TARGET_TAU, reward_gamma=reward_gamma, reward_type=reward_type,
                  test_seeds=test_seeds, episodes_before_train=EPISODES_BEFORE_TRAIN, traffic_density=traffic_density)

    masac.load(model_dir, train_mode=True)
    env.seed = env.config['seed']
    env.unwrapped.seed = env.config['seed']
    eval_rewards = []
    print("train begin")
    ep_rewards = []
    speeds = []
    v_losss = []
    i = 0
    j = 1
    cursh_num = 0

    print("路径", path, "模式", masac.traffic_density, "seeds", env.seed,"车辆数",masac.n_agents)
    sheet.cell(1, 2, "回合奖励")
    sheet.cell(1, 3, "回合损失")
    sheet.cell(1, 4, "回合速度")
    sheet.cell(1, 6, "评价奖励")
    sheet.cell(1, 7, "平均速度")
    sheet.cell(1, 8, "碰撞次数")
    sheet.cell(1, 9, "碰撞率")
    sheet.cell(1, 10, "车头时距")
    sheet.cell(1, 11, "汇入成本")

    while masac.n_episodes < MAX_EPISODES:
        # print(traffic_density)
        masac.interact()  # 交互
        if masac.n_episodes >= EPISODES_BEFORE_TRAIN:
            masac.train()  # 更新
        i = masac.n_episodes - 1
        reward = float(masac.episode_rewards[i])
        speed = float(masac.average_speed[i])

        if masac.episode_done:
            print("回合数：", i, "当前回合奖励：", reward, "平均速度", speed,"车辆数",masac.n_agents)
            # 将当前回合的数据写入Excel
            sheet.cell(i + 2, 1, i)
            sheet.cell(i + 2, 2, reward)
            sheet.cell(i + 2, 4, speed)
            workbook.save(path)

            cursh_num += masac.cursh

        #ep_rewards.append(reward)
        #speeds.append(speed)
        #v_losss.append(masac.v_loss1)
        #if ((i % 200) == 0 or i == 10) and i != 0:
            #qidian = 0
            #now = time.localtime()
            #nowt = time.strftime("%Y-%m-%d/%H.%M.%S", now)
            #print('列入中,当前回合：', i, '时间：', nowt)
            #if i % 200 == 0:
                #qidian = i - 200
            #for k in tqdm(range(qidian, i)):
                #time.sleep(0.1)
                #sheet.cell(k + 2, 1, k)
                # print(rewards[0],rewards[k])
                #sheet.cell(k + 2, 2, ep_rewards[k])
                # sheet.cell(k + 2, 3, v_losss[k])
                #sheet.cell(k + 2, 4, speeds[k])
            workbook.save(path)

        # print("Episode %d, Average Reward %.2f" % (masac.n_episodes +1 , masac.episode_rewards))
        if masac.episode_done and ((masac.n_episodes + 1) % EVAL_INTERVAL == 0):
            j = j + 1
            rewards, _, _, speeds_eval, headway_dis, merging_cost = masac.evaluation(env_eval, dirs['train_videos'],
                                                                                     EVAL_EPISODES)

            rewards_mu, rewards_std = agg_double_list(rewards)
            av_speed = np.mean(speeds_eval)
            av_head_time = np.mean(headway_dis) / speed
            av_merging_cost = np.mean(merging_cost)
            # print(rewards_std,speeds,av_speed,111)
            print('回合', masac.n_episodes + 1, "平均奖励", rewards_mu, "奖励标准差", rewards_std, '平均速度', av_speed,
                  "平均车头时距", av_head_time, "汇入成本", av_merging_cost)
            # print("Episode %d, Average Reward %.2, Averge_speed %.2, Standard deviation reward %.2" % (mappo.n_episodes + 1, rewards_mu,av_speed,rewards_std))
            eval_rewards.append(rewards_mu)
            # save the model
            masac.save(dirs['models'], masac.n_episodes + 1)
            print("eval_reward导入")
            sheet.cell(j, 6, float(rewards_mu))
            # sheet.cell(j, 9, float(rewards_std))
            sheet.cell(j, 7, float(av_speed))
            sheet.cell(j, 8, float(cursh_num))
            sheet.cell(j, 9, float(cursh_num) / 200)
            sheet.cell(j, 10, float(av_head_time))
            sheet.cell(j, 11, float(av_merging_cost))
            workbook.save(path)
            cursh_num = 0
        # np.save(output_dir + '/{}'.format('episode_rewards'), np.array(masac.episode_rewards))
        # np.save(output_dir + '/{}'.format('eval_rewards'), np.array(eval_rewards))
        # np.save(output_dir + '/{}'.format('average_speed'), np.array(masac.average_speed))
    # save the model
    masac.save(dirs['models'], MAX_EPISODES + 2)

    plt.figure()
    plt.plot(eval_rewards)
    plt.xlabel("Episode")
    plt.ylabel("Average Reward")
    plt.legend(["MAsac"])
    plt.show()


def evaluate(args):
    if os.path.exists(args.model_dir):
        model_dir = args.model_dir + '/models/'
    else:
        raise Exception("Sorry, no pretrained models")
    config_dir = args.model_dir + '/configs/configs_sac.ini'
    config = configparser.ConfigParser()
    config.read(config_dir)

    video_dir = args.model_dir + '/eval_videos'

    # model configs
    BATCH_SIZE = config.getint('MODEL_CONFIG', 'BATCH_SIZE')
    MEMORY_CAPACITY = config.getint('MODEL_CONFIG', 'MEMORY_CAPACITY')
    ROLL_OUT_N_STEPS = config.getint('MODEL_CONFIG', 'ROLL_OUT_N_STEPS')
    reward_gamma = config.getfloat('MODEL_CONFIG', 'reward_gamma')
    actor_hidden_size = config.getint('MODEL_CONFIG', 'actor_hidden_size')
    critic_hidden_size = config.getint('MODEL_CONFIG', 'critic_hidden_size')
    MAX_GRAD_NORM = config.getfloat('MODEL_CONFIG', 'MAX_GRAD_NORM')
    ENTROPY_REG = config.getfloat('MODEL_CONFIG', 'ENTROPY_REG')
    reward_type = config.get('MODEL_CONFIG', 'reward_type')
    TARGET_UPDATE_STEPS = config.getint('MODEL_CONFIG', 'TARGET_UPDATE_STEPS')
    TARGET_TAU = config.getfloat('MODEL_CONFIG', 'TARGET_TAU')

    # train configs
    actor_lr = config.getfloat('TRAIN_CONFIG', 'actor_lr')
    critic_lr = config.getfloat('TRAIN_CONFIG', 'critic_lr')
    EPISODES_BEFORE_TRAIN = config.getint('TRAIN_CONFIG', 'EPISODES_BEFORE_TRAIN')
    reward_scale = config.getfloat('TRAIN_CONFIG', 'reward_scale')

    # init env
    env = gym.make('merge-multi-agent-v0')
    env.config['seed'] = config.getint('ENV_CONFIG', 'seed')
    env.config['simulation_frequency'] = config.getint('ENV_CONFIG', 'simulation_frequency')
    env.config['duration'] = config.getint('ENV_CONFIG', 'duration')
    env.config['policy_frequency'] = config.getint('ENV_CONFIG', 'policy_frequency')
    env.config['COLLISION_REWARD'] = config.getint('ENV_CONFIG', 'COLLISION_REWARD')
    env.config['HIGH_SPEED_REWARD'] = config.getint('ENV_CONFIG', 'HIGH_SPEED_REWARD')
    env.config['HEADWAY_COST'] = config.getint('ENV_CONFIG', 'HEADWAY_COST')
    env.config['HEADWAY_TIME'] = config.getfloat('ENV_CONFIG', 'HEADWAY_TIME')
    env.config['MERGING_LANE_COST'] = config.getint('ENV_CONFIG', 'MERGING_LANE_COST')
    env.config['traffic_density'] = config.getint('ENV_CONFIG', 'traffic_density')
    traffic_density = config.getint('ENV_CONFIG', 'traffic_density')
    env.config['action_masking'] = config.getboolean('MODEL_CONFIG', 'action_masking')

    assert env.T % ROLL_OUT_N_STEPS == 0
    state_dim = env.n_s
    action_dim = env.n_a
    test_seeds = args.evaluation_seeds
    seeds = [int(s) for s in test_seeds.split(',')]

    masac = MASAC(env=env, n_agents=len(env.controlled_vehicles), memory_capacity=MEMORY_CAPACITY,
                  state_dim=state_dim, action_dim=action_dim,
                  batch_size=BATCH_SIZE, entropy_reg=ENTROPY_REG,hidden_dim=args.hidden_dim,
                  roll_out_n_steps=ROLL_OUT_N_STEPS,
                  actor_hidden_dim=actor_hidden_size, critic_hidden_dim=critic_hidden_size,
                  actor_lr=actor_lr, critic_lr=critic_lr, reward_scale=reward_scale,
                  tau=TARGET_TAU,
                  reward_gamma=reward_gamma, reward_type=reward_type,
                  max_grad_norm=MAX_GRAD_NORM, test_seeds=test_seeds,
                  episodes_before_train=EPISODES_BEFORE_TRAIN, traffic_density=traffic_density)

    # load the model if exist
    masac.load(model_dir, train_mode=False)
    rewards, _, steps, avg_speeds = masac.evaluation(env, video_dir, len(seeds), is_train=False)


if __name__ == "__main__":
    args = parse_args()

    # train or eval
    if args.option == 'train':
        train(args)
    else:
        evaluate(args)
